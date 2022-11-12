import openpyxl
from openpyxl import load_workbook
import os

# txt转excel
def txt2excel(filepath1,filepath2,txt_sep=','):
    if os.path.isfile(filepath2):
        wb = load_workbook(filepath2)
        ws = wb[wb.sheetnames[0]]
    else:
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(title='Sheet1')
    with open(filepath1,'r') as f:
        ls = f.readlines()
        for l in ls:
            l = l.strip('\n').split(txt_sep)
            l = [a.strip("\"") for a in l]
            ws.append(l)
    wb.save(filepath2)
    wb.close()

def copy_one_row(row):
    l = []
    for r in row:
        l.append(r.value)
    return l
